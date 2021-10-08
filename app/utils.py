from django.core.mail import send_mail
from django.template.loader import render_to_string
import threading

import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment

from .credentials import creds

from googleapiclient.discovery import build
from apiclient import errors
from df2gspread import df2gspread as d2g
# define the scope


# authorize the clientsheet
service_excel = build('sheets', 'v4', credentials=creds)
service_drive = build('drive', 'v2', credentials=creds)


class SheetSnipets(threading.Thread):
    def __init__(self, service_excel):
        self.service_excel = service_excel
        threading.Thread.__init__(self)

    def run(self, data,sheet_id):

        # The ID of the spreadsheet to update.
        # TODO: Update placeholder value.

        spreadsheet_id = sheet_id

        # The A1 notation of a range to search for a logical table of data.
        # Values will be appended after the last row of the table.
        range_ = 'R[1]C[0]'  # TODO: Update placeholder value.

        # How the input data should be interpreted.
        value_input_option = 'RAW'          # TODO: Update placeholder value.

        # How the input data should be inserted.
        insert_data_option = 'INSERT_ROWS'  # TODO: Update placeholder value.

        value_range_body = {
            "values": [  # The data that was read or to be written. This is an array of arrays, the outer array representing all the data and each inner array representing a major dimension. Each item in the inner array corresponds with one cell. For output, empty trailing rows and columns will not be included. For input, supported value types are: bool, string, and double. Null values will be skipped. To set a cell to an empty value, set the string value to an empty string.
                data
            ],
        }
        try :
            request = self.service_excel.spreadsheets().values().append(spreadsheetId=spreadsheet_id, range=range_,
                                                                    valueInputOption=value_input_option, insertDataOption=insert_data_option, body=value_range_body)
            response = request.execute()
        except:
            pass



class SpreadsheetSnippets(threading.Thread):
    def __init__(self, service_excel, service_drive):
        self.service_excel = service_excel
        self.service_drive = service_drive
        threading.Thread.__init__(self)

    def run(self, data, title, approver_email, admin_email, question_type):
        spreadsheetID = self.create(title=title)
        # print('Sheet created :'+spreadsheetID)
        self.addData(data=data, spreadsheetID=spreadsheetID)
        # print('Data Added')
        self.moveData(file_id=spreadsheetID,
                      folder_id='1W1z4j5PHxAUjr4H8tsWJAXb6vS86zAVh')
        # print('Move success')
        self.setPermisionAdmin(spreadsheetID=spreadsheetID, email=admin_email)
        # print('Permission Added for Admin')
        if approver_email != '':
            self.setPermisionApprover(
                spreadsheetID=spreadsheetID, email=approver_email)
        # print('Permission Added for Approver')
        self.styleUpdate(spreadsheetID=spreadsheetID,
                         sheetID=0, question_type=question_type)

    def create(self, title):
        service = self.service_excel
        # [START sheets_create]
        folder_id = '1W1z4j5PHxAUjr4H8tsWJAXb6vS86zAVh'
        spreadsheet = {
            'properties': {
                'title': title
            }
        }
        spreadsheet = service.spreadsheets().create(
            body=spreadsheet, fields='spreadsheetId').execute()
        # [END sheets_create]
        return spreadsheet.get('spreadsheetId')

    def moveData(self, file_id, folder_id):
        service = self.service_drive
        file = service.files().get(fileId=file_id, fields='parents').execute()
        previous_parents = ",".join([parent["id"]
                                    for parent in file.get('parents')])
        # Move the file to the new folder
        file = service.files().update(fileId=file_id, addParents=folder_id,
                                      removeParents=previous_parents, fields='id, parents').execute()

    def addData(self, data, spreadsheetID):
        d2g.upload(data, spreadsheetID, credentials=creds, row_names=True)

    def setPermisionApprover(self, spreadsheetID, email):
        self.insert_permission(spreadsheetID, email, 'user', 'writer')

    def setPermisionAdmin(self, spreadsheetID, email):
        self.insert_permission(spreadsheetID, email, 'user', 'owner')

    def insert_permission(self, file_id, value, perm_type, role):
        service = self.service_drive
        """Insert a new permission.

        Args:
            service: Drive API service instance.
            file_id: ID of the file to insert permission for.
            value: User or group e-mail address, domain name or None for 'default'
                type.
            perm_type: The value 'user', 'group', 'domain' or 'default'.
            role: The value 'owner', 'writer' or 'reader'.
        Returns:
            The inserted permission if successful, None otherwise.
        """
        new_permission = {
            'value': value,
            'type': perm_type,
            'role': role
        }
        try:
            return service.permissions().insert(
                fileId=file_id, body=new_permission).execute()
        except errors.HttpError as error:
            print('An error occurred: %s' % error)
        return None

    def styleUpdate(self, spreadsheetID, sheetID, question_type):
        service = self.service_excel
        requests = []
        if question_type == 'LF':
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheetID,
                            "startRowIndex": 0,
                            "endRowIndex": 2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 3
                            },
                    "mergeType": "MERGE_ROWS"
                }
            })
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheetID,
                            "startRowIndex": 0,
                            "endRowIndex": 2,
                            "startColumnIndex": 3,
                            "endColumnIndex": 5
                            },
                    "mergeType": "MERGE_ROWS"
                }
            })
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheetID,
                            "startRowIndex": 0,
                            "endRowIndex": 2,
                            "startColumnIndex": 5,
                            "endColumnIndex": 7
                            },
                    "mergeType": "MERGE_ROWS"
                }
            })
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheetID,
                            "startRowIndex": 0,
                            "endRowIndex": 2,
                            "startColumnIndex": 7,
                            "endColumnIndex": 9
                            },
                    "mergeType": "MERGE_ROWS"
                }
            })

        elif question_type == 'DQ':
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheetID,
                            "startRowIndex": 2,
                            "endRowIndex": 7,
                            "startColumnIndex": 3,
                            "endColumnIndex": 3
                            },
                    "mergeType": "MERGE_COLUMNS"
                }
            })

        body = {
            'requests': requests
        }
        response = service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheetID,
            body=body).execute()

        # [END sheets_batch_update]
        print(response)


class ExcelGenLF(threading.Thread):

    def __init__(self, data):
        threading.Thread.__init__(self)
        self.data = data

    def run(self):
        df = pd.DataFrame(self.data)

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        ws.delete_cols(1)
        ws.merge_cells('A1:B1')
        ws.merge_cells('A3:B3')

        ws.merge_cells('C1:D1')
        ws.merge_cells('C3:D3')

        ws.merge_cells('E1:F1')
        ws.merge_cells('E3:F3')

        ws.merge_cells('G1:H1')
        ws.merge_cells('G3:H3')

        wb.save("pandas_openpyxl.xlsx")


class ExcelGenDQ(threading.Thread):

    def __init__(self, data):
        threading.Thread.__init__(self)
        self.data = data

    def run(self):
        df = pd.DataFrame(self.data)

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        ws.delete_cols(1)

        ws.merge_cells('C3:C8')

        wb.save("pandas_openpyxl.xlsx")


class EmailThread(threading.Thread):

    def __init__(self, data):
        threading.Thread.__init__(self)
        self.data = data

    def run(self):
        html_message = render_to_string(
            'html_message.html', {'data': self.data}
        )
        send_mail(
            subject=self.data['email_subject'], message=self.data['email_body'], from_email=None, recipient_list=[self.data['to_email']], html_message=html_message)


class Util:
    @staticmethod
    def send_email(data):
        EmailThread(data).start()

    @staticmethod
    def excel_gen_LF(data):
        ExcelGenLF(data).start()

    @staticmethod
    def excel_gen_DQ(data):
        ExcelGenDQ(data).start()

    @staticmethod
    def excel_sheet(service_excel, service_drive, data, title, approver_email, admin_email, question_type, filename):
        SpreadsheetSnippets(service_excel, service_drive).run(
            data, title, approver_email, admin_email, question_type, filename)

    @staticmethod
    def excel_sheet2(service_excel, row,sheet_id):
        SheetSnipets(service_excel).run(data=row,sheet_id=sheet_id)
